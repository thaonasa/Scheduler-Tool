import os
import json
import uuid
import datetime as dt
from io import BytesIO
import re

from flask import Flask, request, render_template_string, send_file, redirect, url_for, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ========== CẤU HÌNH CHUNG ==========
COMPANY_NAME = "Đồng Tiến Bakery"
DATA_PATH = os.path.join(os.path.dirname(__file__), "data", "meeting_schedule.json")  # Lưu JSON
WEEK_DAYS = 6  # Thứ 2 -> Thứ 7

# Bảng màu Chủ trì
CHAIR_COLORS = {
    'TGĐ': '#fcba03',
    'CEO': '#FF9999',
    'COO': '#99CCFF',
    'CFO': '#FF6666',
    'QĐ.XSX': '#FFCC00',
    'QLĐH bánh kem': '#33CC99',
    'TB. KS - TL BGĐ': '#CC99FF',
    'Cố vấn BGĐ': '#FF66CC',
    'TL.BGĐ': '#99CC00',
    'PP kế toán': '#66CCCC',
    'TP Kế Toán NB': '#FF9933',
    'Kế Toán Trưởng': '#339966',
    'Nhân viên QA hệ thống': '#99CCFF',
    'GS.XD DOTICOM': '#CCCC99',
    'PP.NS_ĐT': '#FF99CC',
    'TP.NS_ĐT': '#66CC99',
    'PT kinh doanh': '#3399FF',
    'PP. KD': '#9966FF',
    'QL Product Marketing': '#FF3399',
    'TP. Marketing': '#CC33CC',
    'QL Content Marketing': '#993366',
    'Trưởng Phòng QC': '#66FF99',
    'Phó phòng KHVT': '#FFCC00',
    'IT.PM': '#339999',
    'GS.IT': '#FFCC00',
    'Trưởng BP AI': '#9933CC',
    'Trưởng Ban An Ninh': '#996633',
    'NV ISO': '#CCFF33',
    'NV Phát triển SP': '#669933',
    'Ban HSE': '#66FF00',
    'TBHSE': '#FFCCFF',
    'Trưởng chi nhánh': '#FF6600',
    'HC nhân sự': '#666633',
}

# Danh sách Loại & Phòng họp cho select
CATEGORIES = ["Họp định kỳ", "Họp nội bộ", "Đào tạo", "Phỏng vấn"]
ROOMS = ["Phòng họp 1", "Phòng họp 2", "Phòng họp 3", "Phòng Tổng Giám Đốc"]

# ========== TIỆN ÍCH ==========
def ensure_data_file():
    data_dir = os.path.dirname(DATA_PATH)
    os.makedirs(data_dir, exist_ok=True)
    if not os.path.exists(DATA_PATH):
        with open(DATA_PATH, "w", encoding="utf-8") as f:
            json.dump({"sessions": []}, f, ensure_ascii=False, indent=2)

def load_data():
    ensure_data_file()
    with open(DATA_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def save_data(data):
    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def monday_of_week(any_date: dt.date) -> dt.date:
    return any_date - dt.timedelta(days=any_date.weekday())

def saturday_of_week(any_date: dt.date) -> dt.date:
    return monday_of_week(any_date) + dt.timedelta(days=WEEK_DAYS - 1)

def session_id_from_date(any_date: dt.date) -> str:
    iso_year, iso_week, _ = any_date.isocalendar()
    return f"{iso_year}-W{iso_week:02d}"

def excel_color(hex_color: str) -> str:
    c = hex_color.strip()
    if c.startswith("#"):
        c = c[1:]
    if len(c) == 6:
        return "FF" + c.upper()
    if len(c) == 8:
        return c.upper()
    return "FF000000"

def hhmm_to_minutes(hhmm: str) -> int:
    h, m = map(int, hhmm.split(":"))
    return h * 60 + m

def guess_buoi(start_hhmm: str) -> str:
    return "SÁNG" if hhmm_to_minutes(start_hhmm) < 12 * 60 else "CHIỀU"

def overlap(e1, e2) -> bool:
    if e1["date"] != e2["date"] or e1["session_buoi"] != e2["session_buoi"]:
        return False
    s1, e1t = hhmm_to_minutes(e1["start_time"]), hhmm_to_minutes(e1["end_time"])
    s2, e2t = hhmm_to_minutes(e2["start_time"]), hhmm_to_minutes(e2["end_time"])
    return max(s1, s2) < min(e1t, e2t)

def compute_conflicts(events):
    by_key = {}
    for ev in events:
        key = (ev["date"], ev["session_buoi"])
        by_key.setdefault(key, []).append(ev)

    for key, arr in by_key.items():
        arr.sort(key=lambda x: hhmm_to_minutes(x["start_time"]))
        for i in range(len(arr)):
            arr[i]["conflict"] = False
        for i in range(1, len(arr)):
            if overlap(arr[i - 1], arr[i]):
                arr[i]["conflict"] = True
                arr[i - 1]["conflict"] = True

def compute_attendees_location_conflicts(events):
    by_key = {}
    for ev in events:
        key = (ev["date"], ev["session_buoi"])
        by_key.setdefault(key, []).append(ev)

    for key, arr in by_key.items():
        for i, ev in enumerate(arr):
            ev["attendees_conflict"] = False
            ev["location_conflict"] = False
            ev["chair_conflict"] = False
            for j, other_ev in enumerate(arr):
                if i != j:
                    # Chuẩn hóa danh sách người tham dự để so sánh
                    ev_attendees = set([a.strip() for a in ev.get("attendees", "").split(",") if a.strip()])
                    other_attendees = set([a.strip() for a in other_ev.get("attendees", "").split(",") if a.strip()])
                    same_attendees = ev_attendees.intersection(other_attendees)
                    same_location = ev.get("location") and other_ev.get("location") and ev["location"] == other_ev["location"]
                    time_overlap = overlap(ev, other_ev)

                    # Cảnh báo "Trùng giờ" đã được xử lý trong compute_conflicts
                    # Cảnh báo "Trùng địa điểm" nếu cùng địa điểm và thời gian chồng lấn
                    if time_overlap and same_location:
                        ev["location_conflict"] = True
                        other_ev["location_conflict"] = True
                    # Cảnh báo "Trùng thành phần tham dự" nếu có ít nhất một thành phần chung và thời gian chồng lấn
                    if time_overlap and same_attendees:
                        ev["attendees_conflict"] = True
                        other_ev["attendees_conflict"] = True

def get_or_create_session(data, any_date: dt.date):
    sid = session_id_from_date(any_date)
    for s in data["sessions"]:
        if s["id"] == sid:
            return s
    new_session = {
        "id": sid,
        "week_start": monday_of_week(any_date).isoformat(),
        "week_end": saturday_of_week(any_date).isoformat(),
        "events": []
    }
    data["sessions"].append(new_session)
    save_data(data)
    return new_session

def find_session_by_id(data, sid: str):
    for s in data["sessions"]:
        if s["id"] == sid:
            return s
    return None

def upsert_event(session, payload):
    _id = payload.get("id") or str(uuid.uuid4())
    ev = {
        "id": _id,
        "date": payload["date"],
        "session_buoi": payload.get("buoi") or payload.get("session_buoi") or guess_buoi(payload["start_time"]),
        "start_time": payload["start_time"],
        "end_time": payload["end_time"],
        "title": payload["title"],
        "category": payload.get("category", ""),
        "chair": payload.get("chair", ""),
        "attendees": payload.get("attendees", ""),
        "location": payload.get("location", "")
    }

    if hhmm_to_minutes(ev["start_time"]) >= hhmm_to_minutes(ev["end_time"]):
        raise ValueError("Giờ kết thúc phải lớn hơn giờ bắt đầu.")

    for i, e in enumerate(session["events"]):
        if e["id"] == _id:
            session["events"][i] = ev
            return ev

    session["events"].append(ev)
    return ev

def delete_event(session, event_id: str):
    session["events"] = [e for e in session["events"] if e["id"] != event_id]

# ======= DỮ LIỆU GỘP THEO NGÀY/BUỔI (dùng cho Export & Preview) =======
def build_schedule(session):
    dates = []
    schedule = {}
    week_start = dt.date.fromisoformat(session["week_start"])
    for i in range(WEEK_DAYS):
        date = week_start + dt.timedelta(days=i)
        dates.append(date)
        schedule[date.isoformat()] = {"SÁNG": [], "CHIỀU": []}  # Sử dụng định dạng ISO

    for event in session["events"]:
        date = dt.date.fromisoformat(event["date"])
        if date.isoformat() in schedule:
            schedule[date.isoformat()][event["session_buoi"]].append(event)
        else:
            print(f"Ngày không hợp lệ trong schedule: {event['date']}")

    print(f"Schedule sau khi build: {schedule}")
    return dates, schedule

# ========== XUẤT EXCEL DẠNG BẢNG LỊCH HỌP ==========
def export_session_to_excel(session):
    print(f"Bắt đầu xuất file Excel cho session: {session['id']}")
    wb = Workbook()
    ws = wb.active

    # Kiểm tra và lấy dữ liệu session
    if not session.get("week_start") or not session.get("week_end"):
        print(f"Lỗi: session thiếu week_start hoặc week_end - {session}")
        raise ValueError("Dữ liệu session không hợp lệ")
    
    week_start = dt.date.fromisoformat(session["week_start"])
    week_end = dt.date.fromisoformat(session["week_end"])
    print(f"Tuần: {week_start} -> {week_end}")

    dates, schedule = build_schedule(session)
    print(f"Dates từ build_schedule: {[d.isoformat() for d in dates]}")
    print(f"Schedule từ build_schedule: {schedule}")
    weekdays = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7']

    # Row 1: Tiêu đề
    ws.merge_cells('A1:G1')
    ws['A1'] = f"LỊCH HỌP TUẦN {COMPANY_NAME.upper()}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Row 2: Tuần
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Tuần:  {week_start.strftime('%d/%m/%Y')} -> {week_end.strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Row 3: Header
    ws['A3'] = "BUỔI"  # In hoa và tăng cỡ chữ
    ws['A3'].font = Font(bold=True, size=16)  # Tăng cỡ chữ lên 16
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    date_keys = [d.isoformat() for d in dates]
    for i, (weekday, date) in enumerate(zip(weekdays, dates), 2):
        ws.cell(row=3, column=i, value=f"{weekday}\n({date.strftime('%d.%m.%Y')})")
        ws.cell(row=3, column=i).alignment = Alignment(horizontal="center", wrap_text=True)
        ws.cell(row=3, column=i).font = Font(bold=True)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    start_row = 4
    for buoi in ["SÁNG", "CHIỀU"]:
        max_events = max([len(schedule.get(k, {}).get(buoi, [])) for k in date_keys] + [0])
        if max_events < 1:
            max_events = 1
        block_height = max_events * 2
        end_row = start_row + block_height - 1
        print(f"Buổi {buoi}: max_events={max_events}, block_height={block_height}")

        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        a = ws.cell(row=start_row, column=1, value=buoi)
        a.font = Font(bold=True, size=16)  # Tăng cỡ chữ lên 16
        a.alignment = Alignment(horizontal="center", vertical="center")
        a.border = border

        for col_idx, k in enumerate(date_keys, start=2):
            evs = list(schedule.get(k, {}).get(buoi, []))
            print(f"Ngày {k}, buổi {buoi}: {len(evs)} sự kiện")
            evs.sort(key=lambda e: hhmm_to_minutes(e["start_time"]))

            for r_off in range(max_events):
                header_row = start_row + r_off * 2
                detail_row = header_row + 1
                for rr in (header_row, detail_row):
                    cc = ws.cell(row=rr, column=col_idx)
                    cc.border = border

                if r_off < len(evs):
                    ev = evs[r_off]
                    print(f"Sự kiện {r_off}: {ev}")
                    header_text = f"* {ev['start_time']} - {ev['end_time']}: {ev['title']}\nChủ trì: {ev['chair']}"
                    hcell = ws.cell(row=header_row, column=col_idx, value=header_text)
                    hcell.font = Font(bold=True)
                    hcell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

                    details = []
                    if ev.get("attendees"):
                        details.append(f"- Tham dự: {ev['attendees']}")
                    if ev.get("location"):
                        details.append(f"- Địa điểm: {ev['location']}")
                    if ev.get("category"):
                        details.append(f"- Loại: {ev['category']}")
                    if ev.get("conflict"):
                        details.append("⚠ Trùng giờ")
                    if ev.get("attendees_conflict"):
                        details.append("⚠ Trùng thành phần")
                    if ev.get("location_conflict"):
                        details.append("⚠ Trùng địa điểm")

                    dcell = ws.cell(row=detail_row, column=col_idx, value="\n".join(details))
                    dcell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

                    # Khôi phục tính năng tô màu
                    hexcol = CHAIR_COLORS.get(ev["chair"])
                    if hexcol:
                        fill = PatternFill(start_color=excel_color(hexcol), end_color=excel_color(hexcol), fill_type="solid")
                        hcell.fill = fill
                        dcell.fill = fill

                    # Tự động điều chỉnh chiều cao dòng với hệ số bù
                    h_lines = len(header_text.split('\n')) + 1
                    d_lines = len(details) + 1 if details else 2
                    ws.row_dimensions[header_row].height = max(40, h_lines * 18)
                    ws.row_dimensions[detail_row].height = max(100, d_lines * 18)
                else:
                    hcell = ws.cell(row=header_row, column=col_idx, value="—")
                    hcell.alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=detail_row, column=col_idx, value="")
                    ws.row_dimensions[header_row].height = 40
                    ws.row_dimensions[detail_row].height = 20

        # Thêm hàng trống giữa SÁNG và CHIỀU
        if buoi == "SÁNG":
            start_row = end_row + 2
        else:
            start_row = end_row + 1

    for c in range(1, 8):
        ws.cell(row=3, column=c).border = border

    # Tăng chiều rộng cột để hiển thị nội dung dài
    ws.column_dimensions['A'].width = 15
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 45

    notes_row = start_row + 1
    ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=6)
    ws.cell(row=notes_row, column=1, value="Ghi chú: Các cuộc họp phát sinh TL.BGĐ xin ý kiến BGĐ thống nhất -> HV cập nhật lên phần mềm")
    ws.cell(row=notes_row, column=7, value=f"Đà Nẵng, Ngày {week_end.strftime('%d')} tháng {week_end.strftime('%m')} năm {week_end.strftime('%Y')}").alignment = Alignment(horizontal='center')

    ws.cell(row=notes_row+1, column=2, value="BGĐ KIỂM TRA")
    ws.cell(row=notes_row+1, column=4, value="TP.NS&ĐT Kiểm tra")
    ws.cell(row=notes_row+1, column=6, value="Người lập biểu")

    ws.cell(row=notes_row+4, column=2, value="Phan Thị Yến Tuyết")
    ws.cell(row=notes_row+4, column=4, value="Trần Thị Kim Oanh")
    ws.cell(row=notes_row+4, column=6, value="Trần Thị Mỹ Tân")

    # Lưu file và kiểm tra lỗi
    output = BytesIO()
    try:
        wb.save(output)
        output.seek(0)
        print(f"Xuất file Excel thành công cho session {session['id']}, kích thước: {output.tell()} bytes")
        return output, f"lich_hop_tuan_{session['id']}.xlsx"
    except Exception as e:
        print(f"Lỗi khi xuất file Excel: {e}")
        raise

# ========== XUẤT ICS ==========
def export_session_to_ics(session):
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        f"PRODID:-//{COMPANY_NAME}//Meeting Calendar//VN"
    ]
    for ev in session["events"]:
        date = dt.date.fromisoformat(ev["date"])
        start = dt.datetime.combine(date, dt.time.fromisoformat(ev["start_time"] + ":00"))
        end = dt.datetime.combine(date, dt.time.fromisoformat(ev["end_time"] + ":00"))
        uid = ev["id"]
        title = ev["title"].replace("\n", " ")
        desc = []
        if ev.get("chair"): desc.append(f"Chu tri: {ev['chair']}")
        if ev.get("attendees"): desc.append(f"Tham du: {ev['attendees']}")
        if ev.get("category"): desc.append(f"Loai: {ev['category']}")
        description = "\\n".join(desc)
        def fmt(dtobj): return dtobj.strftime("%Y%m%dT%H%M%SZ")
        lines += [
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{fmt(dt.datetime.utcnow())}",
            f"DTSTART:{fmt(start)}",
            f"DTEND:{fmt(end)}",
            f"SUMMARY:{title}",
            f"DESCRIPTION:{description}",
            f"LOCATION:{ev.get('location','')}",
            "END:VEVENT"
        ]
    lines.append("END:VCALENDAR")
    ics_bytes = "\r\n".join(lines).encode("utf-8")
    return BytesIO(ics_bytes), f"lich_hop_tuan_{session['id']}.ics"

# ========== IMPORT TỪ EXCEL ==========
def import_from_excel(file, target_date: dt.date):
    wb = load_workbook(file)
    ws = wb.active

    # Lấy ngày từ tuần mục tiêu
    data = load_data()
    target_session = get_or_create_session(data, target_date)
    target_week_start = dt.date.fromisoformat(target_session["week_start"])
    week_days = [target_week_start + dt.timedelta(days=i) for i in range(6)]  # Thứ 2 đến Thứ 7

    # Xác định buoi và nội dung
    contents = {i+2: [] for i in range(6)}  # Danh sách sự kiện cho mỗi cột (B đến G)
    current_buoi = None
    for row in range(4, ws.max_row + 1):
        a_cell = ws.cell(row=row, column=1).value
        if a_cell:
            a_cell = a_cell.strip()
            if a_cell in ['Ghi chú:', 'BGĐ KIỂM TRA']:  # Dừng khi đến ghi chú
                break
            if a_cell in ['SÁNG', 'CHIỀU']:
                current_buoi = a_cell
                continue
        if current_buoi is None:
            continue

        # Nối nội dung cho từng cột
        for col in range(2, 8):  # Cột B đến G
            cell = ws.cell(row=row, column=col).value
            if cell:
                contents[col].append((current_buoi, cell))

    # Parse và thêm sự kiện
    imported_count = 0
    for col, day in enumerate(week_days, start=2):
        for buoi, content in contents.get(col, []):
            if not content:
                continue
            parsed_events = parse_cell(content)
            for parsed in parsed_events:
                start_time_minutes = hhmm_to_minutes(parsed['start_time'])
                session_buoi = buoi
                if start_time_minutes >= 12 * 60 and buoi == "SÁNG":
                    session_buoi = "CHIỀU"
                elif start_time_minutes < 12 * 60 and buoi == "CHIỀU":
                    session_buoi = "SÁNG"

                payload = {
                    "id": str(uuid.uuid4()),
                    "date": day.isoformat(),
                    "session_buoi": session_buoi,
                    "start_time": parsed['start_time'],
                    "end_time": parsed['end_time'],
                    "title": parsed['title'],
                    "category": parsed['category'],
                    "chair": parsed['chair'],
                    "attendees": parsed['attendees'],
                    "location": parsed['location']
                }
                try:
                    event = upsert_event(target_session, payload)
                    imported_count += 1
                    print(f"Đã thêm sự kiện: {event['title']} - {event['date']} {event['session_buoi']} {event['start_time']}")
                except ValueError as e:
                    print(f"Lỗi khi thêm sự kiện: {e} - Payload: {payload}")

    save_data(data)
    print(f"Đã import thành công {imported_count} sự kiện.")
    return target_session["id"]

def parse_cell(cell_content):
    if not cell_content:
        return []

    # Tách các sự kiện dựa trên dấu * 
    events = re.split(r'\*(\d{2}h\d{2} - \d{2}h\d{2}):', cell_content.strip())
    parsed_events = []

    i = 0
    while i < len(events):
        if i % 2 == 0:
            i += 1
            continue
        time_str = events[i].strip()
        details = events[i+1].strip() if i+1 < len(events) else ''
        i += 2

        # Tách giờ bắt đầu và kết thúc
        time_parts = re.match(r'(\d{2}h\d{2}) - (\d{2}h\d{2})', time_str)
        if time_parts:
            start_time = time_parts.group(1).replace('h', ':')
            end_time = time_parts.group(2).replace('h', ':')
        else:
            print(f"Không thể parse thời gian: {time_str}")
            continue

        # Tách tiêu đề (dòng đầu details)
        lines = details.split('\n')
        title = lines[0].strip() if lines else ''

        # Các dòng tiếp: Chủ trì, Tham dự, Địa điểm
        chair = ''
        attendees = ''
        location = ''
        category = ''
        for line in lines[1:]:
            line = line.strip()
            if line.startswith('Chủ trì:'):
                chair = line.replace('Chủ trì:', '').strip()
            elif line.startswith('- Tham dự:') or line.startswith('-  Tham dự:'):
                attendees = line.replace('- Tham dự:', '').replace('-  Tham dự:', '').strip()
            elif line.startswith('- Địa điểm:'):
                location = line.replace('- Địa điểm:', '').strip()
            elif line.startswith('- Loại:'):
                category = line.replace('- Loại:', '').strip()

        parsed_events.append({
            'start_time': start_time,
            'end_time': end_time,
            'title': title,
            'chair': chair,
            'attendees': attendees,
            'location': location,
            'category': category
        })

    return parsed_events

def build_schedule(session):
    dates = []
    schedule = {}
    week_start = dt.date.fromisoformat(session["week_start"])
    for i in range(6):
        date = week_start + dt.timedelta(days=i)
        dates.append(date)
        schedule[date.isoformat()] = {"SÁNG": [], "CHIỀU": []}  # Sử dụng định dạng ISO cho key

    for event in session["events"]:
        date = dt.date.fromisoformat(event["date"])
        if date.isoformat() in schedule:
            schedule[date.isoformat()][event["session_buoi"]].append(event)
        else:
            print(f"Ngày không hợp lệ trong schedule: {event['date']}")

    print(f"Schedule sau khi build: {schedule}")
    return dates, schedule


# ========== SAO CHÉP TUẦN ==========
def copy_week_to_another(data, source_session_id, target_date: dt.date):
    source_session = find_session_by_id(data, source_session_id)
    if not source_session:
        raise ValueError("Không tìm thấy tuần nguồn.")
    
    target_session = get_or_create_session(data, target_date)
    target_week_start = dt.date.fromisoformat(target_session["week_start"])

    for event in source_session["events"]:
        event_date = dt.date.fromisoformat(event["date"])
        day_diff = (event_date - monday_of_week(event_date)).days
        adjusted_date = target_week_start + dt.timedelta(days=day_diff)
        
        payload = {
            "id": str(uuid.uuid4()),  # Tạo ID mới cho sự kiện sao chép
            "date": adjusted_date.isoformat(),
            "session_buoi": event["session_buoi"],  # Đảm bảo sử dụng 'session_buoi' từ event
            "start_time": event["start_time"],
            "end_time": event["end_time"],
            "title": event["title"],
            "category": event.get("category", ""),
            "chair": event["chair"],
            "attendees": event.get("attendees", ""),
            "location": event.get("location", "")
        }
        try:
            upsert_event(target_session, payload)
        except ValueError:
            continue
    
    save_data(data)
    return target_session["id"]

# ========== ROUTES ==========
@app.route("/")
def home():
    data = load_data()
    qdate = request.args.get("date")
    today = dt.date.today() if not qdate else dt.date.fromisoformat(qdate)
    sess = get_or_create_session(data, today)

    sessions_sorted = sorted(data["sessions"], key=lambda s: s["week_start"], reverse=True)

    q = request.args.get("q", "").strip().lower()
    events = list(sess["events"])
    if q:
        events = [e for e in events if q in json.dumps(e, ensure_ascii=False).lower()]

    # Cảnh báo xung đột
    compute_conflicts(events)
    compute_attendees_location_conflicts(events)

    # >>> NEW: dữ liệu cho tab "Lịch"
    dates, schedule = build_schedule(sess)
    weekdays = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7']

    return render_template_string(
        TEMPLATE_INDEX,
        company=COMPANY_NAME,
        chair_colors=CHAIR_COLORS,
        categories=CATEGORIES,
        rooms=ROOMS,
        session=sess,
        sessions=sessions_sorted,
        events=sorted(events, key=lambda x: (x["date"], x["session_buoi"], x["start_time"])),
        week_start=dt.date.fromisoformat(sess["week_start"]),
        week_end=dt.date.fromisoformat(sess["week_end"]),
        today=today,
        q=q,
        import_error=None,
        # >>> NEW:
        dates=dates,
        schedule=schedule,
        weekdays=weekdays
    )


@app.route("/preview/<session_id>")
def preview(session_id):
    data = load_data()
    sess = find_session_by_id(data, session_id)
    if not sess:
        return "Không tìm thấy session", 404
    dates, schedule = build_schedule(sess)
    weekdays = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7']
    return render_template_string(
        TEMPLATE_PREVIEW,
        company=COMPANY_NAME,
        chair_colors=CHAIR_COLORS,
        session=sess,
        dates=dates,
        schedule=schedule,
        weekdays=weekdays
    )

@app.route("/sessions")
def list_sessions():
    data = load_data()
    sessions_sorted = sorted(data["sessions"], key=lambda s: s["week_start"], reverse=True)
    return jsonify(sessions_sorted)

@app.route("/switch-session", methods=["POST"])
def switch_session():
    date_str = request.form.get("any_date")
    if not date_str:
        return redirect(url_for("home"))
    return redirect(url_for("home", date=date_str))

@app.route("/event", methods=["POST"])
def add_or_update_event():
    data = load_data()
    date_str = request.form["date"]
    buoi = request.form.get("buoi") or guess_buoi(request.form["start_time"])
    sess = get_or_create_session(data, dt.date.fromisoformat(date_str))

    payload = {
    "id": request.form.get("id", ""),
    "date": date_str,
    "buoi": buoi,
    "start_time": request.form["start_time"],
    "end_time": request.form["end_time"],
    "title": request.form["title"],
    "category": request.form.get("category", ""),
    "chair": request.form["chair"],
    "attendees": ", ".join(request.form.getlist('attendees')) if request.form.getlist('attendees') else "",
    "location": request.form.get("location", "")
}
    try:
        upsert_event(sess, payload)
        save_data(data)
        return redirect(url_for("home", date=date_str))
    except ValueError as e:
        return f"Lỗi: {e}", 400

@app.route("/event/<session_id>/<event_id>/delete", methods=["POST"])
def remove_event(session_id, event_id):
    data = load_data()
    sess = find_session_by_id(data, session_id)
    if not sess:
        return "Không tìm thấy session", 404
    delete_event(sess, event_id)
    save_data(data)
    return redirect(url_for("home", date=sess["week_start"]))

@app.route("/event/<session_id>/clear", methods=["POST"])
def clear_session(session_id):
    data = load_data()
    sess = find_session_by_id(data, session_id)
    if not sess:
        return "Không tìm thấy session", 404
    sess["events"] = []
    save_data(data)
    return redirect(url_for("home", date=sess["week_start"]))

@app.route("/export/<session_id>/excel", methods=["POST"])
def export_excel(session_id):
    data = load_data()
    sess = find_session_by_id(data, session_id)
    if not sess:
        return "Không tìm thấy session", 404
    try:
        output, fname = export_session_to_excel(sess)
        return send_file(
            output,
            as_attachment=True,
            download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Lỗi khi gửi file Excel: {e}")
        return f"Lỗi khi xuất file: {str(e)}", 500

@app.route("/export/<session_id>/ics", methods=["POST"])
def export_ics(session_id):
    data = load_data()
    sess = find_session_by_id(data, session_id)
    if not sess:
        return "Không tìm thấy session", 404
    output, fname = export_session_to_ics(sess)
    return send_file(output,
                     as_attachment=True,
                     download_name=fname,
                     mimetype="text/calendar")

@app.route("/backup/json", methods=["GET"])
def backup_json():
    ensure_data_file()
    return send_file(DATA_PATH, as_attachment=True, download_name="meeting_schedule_backup.json")

@app.route("/import", methods=["POST"])
def import_data():
    data = load_data()
    import_error = None
    if 'file' not in request.files:
        import_error = "Vui lòng chọn một file để tải lên."
    else:
        file = request.files['file']
        if file.filename == '':
            import_error = "Vui lòng chọn một file để tải lên."
        elif not file.filename.endswith('.xlsx'):
            import_error = "Chỉ chấp nhận file Excel (.xlsx)."
        else:
            target_date = dt.date.fromisoformat(request.form.get("target_date", dt.date.today().isoformat()))
            session_id = import_from_excel(file, target_date)
            return redirect(url_for("home", date=target_date.isoformat()))

    qdate = request.args.get("date")
    today = dt.date.today() if not qdate else dt.date.fromisoformat(qdate)
    sess = get_or_create_session(data, today)
    sessions_sorted = sorted(data["sessions"], key=lambda s: s["week_start"], reverse=True)
    q = request.args.get("q", "").strip().lower()
    events = list(sess["events"])
    if q:
        events = [e for e in events if q in json.dumps(e, ensure_ascii=False).lower()]
    compute_conflicts(events)
    compute_attendees_location_conflicts(events)

    return render_template_string(
        TEMPLATE_INDEX,
        company=COMPANY_NAME,
        chair_colors=CHAIR_COLORS,
        categories=CATEGORIES,
        rooms=ROOMS,
        session=sess,
        sessions=sessions_sorted,
        events=sorted(events, key=lambda x: (x["date"], x["session_buoi"], x["start_time"])),
        week_start=dt.date.fromisoformat(sess["week_start"]),
        week_end=dt.date.fromisoformat(sess["week_end"]),
        today=today,
        q=q, 
        import_error=import_error
    )


@app.route("/copy-week", methods=["POST"])
def copy_week():
    data = load_data()
    source_session_id = request.form.get("source_session_id")
    target_date = dt.date.fromisoformat(request.form.get("target_date", dt.date.today().isoformat()))
    
    try:
        target_session_id = copy_week_to_another(data, source_session_id, target_date)
        return redirect(url_for("home", date=target_date.isoformat()))
    except ValueError as e:
        import_error = str(e)
        qdate = request.args.get("date")
        today = dt.date.today() if not qdate else dt.date.fromisoformat(qdate)
        sess = get_or_create_session(data, today)
        sessions_sorted = sorted(data["sessions"], key=lambda s: s["week_start"], reverse=True)
        q = request.args.get("q", "").strip().lower()
        events = list(sess["events"])
        if q:
            events = [e for e in events if q in json.dumps(e, ensure_ascii=False).lower()]
        compute_conflicts(events)
        compute_attendees_location_conflicts(events)

        return render_template_string(
            TEMPLATE_INDEX,
            company=COMPANY_NAME,
            chair_colors=CHAIR_COLORS,
            categories=CATEGORIES,
            rooms=ROOMS,
            session=sess,
            sessions=sessions_sorted,
            events=sorted(events, key=lambda x: (x["date"], x["session_buoi"], x["start_time"])),
            week_start=dt.date.fromisoformat(sess["week_start"]),
            week_end=dt.date.fromisoformat(sess["week_end"]),
            today=today,
            q=q,
            import_error=import_error
        )

# ========== TEMPLATES ==========
TEMPLATE_INDEX = """
<!doctype html>
<html lang="vi">
<head>
  <meta charset="utf-8">
  <title>Lịch Họp – {{ company }}</title>
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <style>
    :root{
      --bg:#f5f7fb; --surface:#ffffff; --text:#1f2937; --muted:#6b7280;
      --border:#e5e7eb; --primary:#2563eb; --danger:#ef4444; --warn:#b45309;
    }
    *{box-sizing:border-box}
    html,body{height:100%}
    body{margin:0;background:var(--bg);color:var(--text);
      font:14px/1.45 ui-sans-serif,system-ui,-apple-system,Segoe UI,Roboto,Helvetica,Arial}
    a{color:inherit;text-decoration:none}
    button,input,select{font:inherit}

    /* =========== LAYOUT =========== */
    .app{min-height:100vh;display:grid;grid-template-rows:auto 1fr auto}
    .header{background:var(--surface);border-bottom:1px solid var(--border);
      display:flex;align-items:center;gap:16px;padding:10px 18px;position:sticky;top:0;z-index:30}
    .brand{display:flex;align-items:center;gap:10px;font-weight:800}
    .brand .logo-wrap{width:28px;height:28px;border-radius:6px;overflow:hidden;display:grid;place-items:center;background:#fff}
    .brand .logo{width:28px;height:28px;object-fit:contain;display:block}
    .logo-fallback{width:28px;height:28px;border-radius:6px;background:var(--primary);display:none;place-items:center;color:#fff}

    .nav{margin-left:auto;display:flex;gap:12px;align-items:center}
    .nav a,.nav button{border:1px solid var(--border);background:var(--surface);padding:8px 10px;border-radius:999px;cursor:pointer}
    .nav a.primary{background:var(--primary);border-color:transparent;color:#fff}

    .main{display:grid;grid-template-columns:290px 1fr;gap:16px;padding:16px}
    @media (max-width: 980px){ .main{grid-template-columns:1fr} }

    .card{background:var(--surface);border:1px solid var(--border);border-radius:12px;box-shadow:0 1px 2px rgba(0,0,0,.04)}
    .card h2{margin:0;padding:14px 16px;border-bottom:1px solid var(--border);font-size:16px;background:#fafafa}
    .card .content{padding:14px 16px}

    .row{display:flex;gap:10px;flex-wrap:wrap}
    .row>*{flex:1}
    input,select,button{border:1px solid var(--border);border-radius:8px;padding:10px 12px;background:#fff}
    button.primary{background:var(--primary);border-color:transparent;color:#fff}
    button.danger{background:var(--danger);border-color:transparent;color:#fff}
    .muted{color:var(--muted)} .nowrap{white-space:nowrap}

    .tag{display:inline-flex;align-items:center;gap:6px;padding:4px 8px;border:1px solid var(--border);border-radius:999px;background:#fff}
    .dot{width:12px;height:12px;border-radius:999px;border:1px solid var(--border)}

    .grid2{display:grid;grid-template-columns:1fr 1fr;gap:10px}
    .grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}
    @media (max-width: 980px){ .grid2,.grid3{grid-template-columns:1fr} }

    /* tabs */
    .tabs{display:flex;gap:8px;padding:0 16px 10px}
    .tab{padding:8px 12px;border:1px solid var(--border);background:#fff;border-radius:999px;cursor:pointer}
    .tab.active{background:var(--primary);color:#fff;border-color:transparent}
    .view{display:none}.view.active{display:block}

    /* calendar */
    .cal{border:1px solid var(--border);border-radius:12px;overflow:hidden;background:#fff}
    .cal-head{display:grid;grid-template-columns:120px repeat(6,1fr);background:#f9fafb;border-bottom:1px solid var(--border)}
    .cal-head>div{padding:10px 12px;text-align:center;font-weight:600}
    .cal-row{display:grid;grid-template-columns:120px repeat(6,1fr);border-bottom:1px solid #f3f4f6}
    .cal-buoi{background:#f9fafb;padding:10px 12px;text-align:center;font-weight:700;border-right:1px solid var(--border)}
    .cal-cell{padding:10px;min-height:190px;border-right:1px solid #f3f4f6}
    @media (max-width:1100px){ .cal-head,.cal-row{grid-template-columns:90px repeat(6,1fr)} }

    /* event card */
    .ev{position:relative;padding:8px 8px 44px;border-radius:10px;background:#f3f4f6;margin-bottom:10px;box-shadow:inset 0 0 0 1px rgba(0,0,0,.05)}
    .ev .tt{font-weight:700}
    .warn{color:var(--warn);font-weight:700}
    .ev .actions{position:absolute;right:8px;bottom:8px;display:flex;gap:6px}
    .ev .actions button{padding:6px 8px;border:1px solid var(--border);border-radius:6px;background:#fff;cursor:pointer}
    .ev .actions .danger{background:var(--danger);color:#fff;border-color:transparent}

    /* attendees checkboxes */
    .checkbox-wrap{margin-top:8px}
    .checkbox-group{display:flex;flex-wrap:wrap;gap:12px;margin-top:6px}
    .checkbox-group label{display:flex;align-items:center;gap:6px;cursor:pointer;user-select:none}

    /* tables */
    table{width:100%;border-collapse:collapse;font-size:14px}
    th,td{padding:8px;border-bottom:1px solid #eee;vertical-align:top;text-align:left}

    /* footer */
    .footer{background:#fff;border-top:1px solid var(--border);padding:16px 18px;display:grid;gap:10px;justify-items:center}
    .footer .legend{display:flex;flex-wrap:wrap;gap:8px;justify-content:center}
    .footer .copy{color:var(--muted);font-size:13px}
  </style>
</head>
<body>
<div class="app">

  <!-- ===== HEADER ===== -->
  <header class="header">
    <a class="brand" href="/">
      <span class="logo-wrap">
        <img class="logo" src="/static/logo.png" alt="Logo"
             onerror="this.style.display='none'; this.parentElement.nextElementSibling.style.display='grid';">
      </span>
      <span class="logo-fallback">🏢</span>
      <span>{{ company }}</span>
    </a>
    <div class="nav">
      <a href="/" class="primary">🏠 Trang chủ</a>
      <form method="post" action="/export/{{ session.id }}/excel" style="display:inline">
        <button type="submit">📤 Export Excel</button>
      </form>
      <form method="post" action="/export/{{ session.id }}/ics" style="display:inline">
        <button type="submit">📆 Export ICS</button>
      </form>
      <a href="/backup/json">🗄️ Backup JSON</a>
    </div>
  </header>

  <!-- ===== MAIN ===== -->
  <main class="main">

    <!-- ===== SIDEBAR ===== -->
    <aside class="card">
      <h2>Tuần &amp; Tính năng</h2>
      <div class="content">
        <!-- Mở tuần -->
        <form method="post" action="/switch-session" class="row" style="margin-bottom:12px">
          <div>
            <label class="muted">Chọn bất kỳ ngày trong tuần</label>
            <input type="date" name="any_date" value="{{ today.isoformat() }}">
          </div>
          <div class="nowrap" style="align-self:flex-end">
            <button class="primary" type="submit">Mở tuần</button>
          </div>
        </form>

        <div class="muted" style="margin:8px 0 12px">
          Tuần hiện tại: <b>{{ week_start.strftime('%d/%m/%Y') }}</b> → <b>{{ week_end.strftime('%d/%m/%Y') }}</b>
        </div>

        <!-- Tìm kiếm -->
        <form method="get" action="/" class="row" style="margin-top:6px">
          <input type="hidden" name="date" value="{{ week_start.isoformat() }}">
          <input type="text" name="q" placeholder="Tìm kiếm..." value="{{ q }}">
          <button type="submit">Lọc</button>
        </form>

        <hr style="margin:14px 0">

        <!-- Xoá toàn tuần -->
        <div class="row">
          <form method="post" action="/event/{{ session.id }}/clear" onsubmit="return confirm('Xoá toàn bộ sự kiện của tuần này?')">
            <button class="danger" type="submit">🗑️ Xoá toàn tuần</button>
          </form>
        </div>

        <!-- Import từ Excel -->
        <hr style="margin:14px 0">
        <form method="post" action="/import" enctype="multipart/form-data" class="row">
          <input type="date" name="target_date" value="{{ today.isoformat() }}">
          <input type="file" name="file" accept=".xlsx">
          <button class="primary" type="submit">📥 Import từ Excel</button>
        </form>

        <!-- Sao chép tuần -->
        <hr style="margin:14px 0">
        <form method="post" action="/copy-week" class="row">
          <select name="source_session_id">
            {% for s in sessions %}
              <option value="{{ s.id }}">{{ s.id }} ({{ s.week_start }} → {{ s.week_end }})</option>
            {% endfor %}
          </select>
          <input type="date" name="target_date" value="{{ today.isoformat() }}">
          <button class="primary" type="submit">📑 Sao chép tuần</button>
        </form>

        {% if import_error %}
        <div style="margin-top:12px;color:#ef4444;padding:8px;border:1px solid #fee2e2;border-radius:8px">
          {{ import_error }}
        </div>
        {% endif %}

        <!-- Các tuần gần đây -->
        <hr style="margin:14px 0">
        <div>
          <div class="muted" style="margin-bottom:6px">Các tuần gần đây</div>
          <div style="max-height:260px;overflow:auto">
            <table>
              <thead><tr><th>Tuần</th><th class="nowrap">Mở</th></tr></thead>
              <tbody>
                {% for s in sessions %}
                <tr>
                  <td><b>{{ s.id }}</b><br><span class="muted">{{ s.week_start }} → {{ s.week_end }}</span></td>
                  <td class="nowrap"><a href="/?date={{ s.week_start }}"><button type="button">Xem</button></a></td>
                </tr>
                {% endfor %}
              </tbody>
            </table>
          </div>
        </div>
      </div>
    </aside>

    <!-- ===== CONTENT ===== -->
    <section class="card">
      <h2>Thêm/Chỉnh sửa sự kiện</h2>
      <div class="content">
        <form id="event-form" method="post" action="/event">
          <input type="hidden" name="id" id="fld-id">

          <div class="grid3">
            <div>
              <label>Ngày</label>
              <input type="date" name="date" id="fld-date" value="{{ week_start.isoformat() }}" required>
            </div>
            <div>
              <label>Buổi</label>
              <select name="buoi" id="fld-buoi">
                <option value="">(Tự nhận diện theo giờ)</option>
                <option value="SÁNG">SÁNG</option>
                <option value="CHIỀU">CHIỀU</option>
              </select>
            </div>
            <div>
              <label>Loại</label>
              <input type="hidden" name="category" id="fld-category">
              <select id="fld-category-select">
                {% for c in categories %}<option value="{{ c }}">{{ c }}</option>{% endfor %}
                <option value="__OTHER__">Khác…</option>
              </select>
              <input type="text" id="fld-category-other" placeholder="Nhập loại khác" style="display:none;margin-top:6px">
            </div>
          </div>

          <div class="grid3" style="margin-top:8px">
            <div>
              <label>Giờ bắt đầu</label>
              <input type="time" name="start_time" id="fld-start" required>
            </div>
            <div>
              <label>Giờ kết thúc</label>
              <input type="time" name="end_time" id="fld-end" required>
            </div>
            <div>
              <label>Chủ trì</label>
              <select name="chair" id="fld-chair" required>
                {% for chair,color in chair_colors.items() %}<option value="{{ chair }}">{{ chair }}</option>{% endfor %}
              </select>
            </div>
          </div>

          <div class="grid2" style="margin-top:8px">
            <div>
              <label>Tên họp</label>
              <input type="text" name="title" id="fld-title" placeholder="VD: Họp giao ban tuần" required>
            </div>
            <div>
              <label>Địa điểm</label>
              <input type="hidden" name="location" id="fld-location">
              <select id="fld-location-select">
                {% for r in rooms %}<option value="{{ r }}">{{ r }}</option>{% endfor %}
                <option value="__OTHER__">Khác…</option>
              </select>
              <input type="text" id="fld-location-other" placeholder="Nhập địa điểm" style="display:none;margin-top:6px">
            </div>
          </div>

          <div class="checkbox-wrap">
            <label>Thành phần tham dự</label>
            <div class="checkbox-group">
              {% for chair in chair_colors.keys() %}
              <label><input type="checkbox" name="attendees" value="{{ chair }}"> {{ chair }}</label>
              {% endfor %}
            </div>
          </div>

          <div class="row" style="margin-top:12px">
            <button class="primary" type="submit">💾 Lưu sự kiện</button>
            <button type="reset" onclick="document.getElementById('fld-id').value=''">🧹 Xoá nhập</button>
          </div>
        </form>
      </div>

      <!-- TABS -->
      <div class="tabs">
        <button class="tab active" data-tab="view-calendar">📅 Lịch</button>
        <button class="tab" data-tab="view-table">📄 Danh sách</button>
      </div>

      <!-- ===== VIEW: CALENDAR ===== -->
      <div id="view-calendar" class="view active">
        <div class="content">
          <div class="row" style="align-items:center;margin-bottom:10px">
            <label style="display:flex;align-items:center;gap:6px"><input type="checkbox" id="only-conflicts"> Chỉ hiển thị sự kiện có cảnh báo</label>
          </div>

          <div class="cal">
            <div class="cal-head">
              <div>Buổi</div>
              {% for d in dates %}
                <div>{{ weekdays[loop.index0] }}<br><span class="muted">({{ d.strftime('%d.%m.%Y') }})</span></div>
              {% endfor %}
            </div>

            {% for buoi in ['SÁNG','CHIỀU'] %}
            <div class="cal-row">
              <div class="cal-buoi">{{ buoi }}</div>
              {% for d in dates %}
                {% set key = d.isoformat() %}
                <div class="cal-cell">
                  {% for ev in schedule.get(key, {}).get(buoi, []) %}
                    {% set bg = chair_colors.get(ev.chair, '#f3f4f6') %}
                    <div class="ev"
                         style="background:{{ bg }}"
                         data-id="{{ ev.id }}"
                         data-date="{{ ev.date }}"
                         data-buoi="{{ ev.session_buoi }}"
                         data-start="{{ ev.start_time }}"
                         data-end="{{ ev.end_time }}"
                         data-title="{{ ev.title|e }}"
                         data-chair="{{ ev.chair }}"
                         data-attendees="{{ ev.attendees|e }}"
                         data-location="{{ ev.location|e }}"
                         data-category="{{ ev.category|e }}"
                         data-has-conflict="{{ '1' if (ev.conflict or ev.attendees_conflict or ev.location_conflict) else '0' }}">
                      <div class="tt">• {{ ev.start_time }}–{{ ev.end_time }}: {{ ev.title }}</div>
                      <div>Chủ trì: <b>{{ ev.chair }}</b></div>
                      {% if ev.attendees %}<div>- Thành phần tham dự: {{ ev.attendees }} {% if ev.attendees_conflict %}<span class="warn">⚠ Trùng thành phần</span>{% endif %}</div>{% endif %}
                      {% if ev.location %}<div>- Địa điểm: {{ ev.location }} {% if ev.location_conflict %}<span class="warn">⚠ Trùng địa điểm</span>{% endif %}</div>{% endif %}
                      {% if ev.category %}<div>- Loại: {{ ev.category }}</div>{% endif %}
                      {% if ev.conflict %}<div class="warn">⚠ Trùng giờ</div>{% endif %}

                      <div class="actions">
                        <button type="button" onclick="editEventFromCard(this)">Sửa</button>
                        <form method="post" action="/event/{{ session.id }}/{{ ev.id }}/delete" onsubmit="return confirm('Xoá sự kiện này?')">
                          <button class="danger" type="submit">Xoá</button>
                        </form>
                      </div>
                    </div>
                  {% else %}
                    <div class="muted" style="font-style:italic">—</div>
                  {% endfor %}
                </div>
              {% endfor %}
            </div>
            {% endfor %}
          </div>
        </div>
      </div>

      <!-- ===== VIEW: TABLE ===== -->
      <div id="view-table" class="view">
        <div class="content">
          {% if events %}
          <table>
            <thead>
              <tr>
                <th class="nowrap">Ngày</th><th>Buổi</th><th>Giờ</th><th>Tiêu đề</th>
                <th>Chủ trì</th><th>Thành phần tham dự</th><th>Địa điểm</th><th>Cảnh báo</th><th></th>
              </tr>
            </thead>
            <tbody>
              {% for ev in events %}
              <tr
                data-id="{{ ev.id }}" data-date="{{ ev.date }}" data-buoi="{{ ev.session_buoi }}"
                data-start="{{ ev.start_time }}" data-end="{{ ev.end_time }}"
                data-title="{{ ev.title|e }}" data-chair="{{ ev.chair }}"
                data-attendees="{{ ev.attendees|e }}" data-location="{{ ev.location|e }}"
                data-category="{{ ev.category|e }}">
                <td class="nowrap">{{ ev.date }}</td>
                <td>{{ ev.session_buoi }}</td>
                <td class="nowrap">{{ ev.start_time }}–{{ ev.end_time }} {% if ev.conflict %}<span class="warn">⚠ Trùng giờ</span>{% endif %}</td>
                <td><div style="font-weight:600">{{ ev.title }}</div>{% if ev.category %}<div class="muted">Loại: {{ ev.category }}</div>{% endif %}</td>
                <td>{{ ev.chair }}</td>
                <td>{{ ev.attendees }} {% if ev.attendees_conflict %}<span class="warn">⚠ Trùng thành phần</span>{% endif %}</td>
                <td>{{ ev.location }} {% if ev.location_conflict %}<span class="warn">⚠ Trùng địa điểm</span>{% endif %}</td>
                <td>
                  {% if ev.conflict %}<div class="warn">⚠ Trùng giờ</div>{% endif %}
                  {% if ev.attendees_conflict %}<div class="warn">⚠ Trùng thành phần</div>{% endif %}
                  {% if ev.location_conflict %}<div class="warn">⚠ Trùng địa điểm</div>{% endif %}
                </td>
                <td class="nowrap">
                  <button type="button" onclick="editEvent(this)">Sửa</button>
                  <form method="post" action="/event/{{ session.id }}/{{ ev.id }}/delete" style="display:inline" onsubmit="return confirm('Xoá sự kiện này?')">
                    <button class="danger" type="submit">Xoá</button>
                  </form>
                </td>
              </tr>
              {% endfor %}
            </tbody>
          </table>
          {% else %}
          <div class="muted">Chưa có sự kiện nào trong tuần này.</div>
          {% endif %}
        </div>
      </div>

    </section>
  </main>

  <!-- ===== FOOTER ===== -->
  <footer class="footer">
    <div class="legend">
      {% for chair, color in chair_colors.items() %}
        <span class="tag"><span class="dot" style="background: {{ color }}"></span>{{ chair }}</span>
      {% endfor %}
    </div>
    <div class="copy">© {{ week_end.strftime('%Y') }} {{ company }} — Hệ thống lịch họp nội bộ</div>
  </footer>
</div>

<script>
  // Tabs
  document.querySelectorAll('.tab').forEach(b=>{
    b.addEventListener('click',()=>{
      document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));
      b.classList.add('active');
      const id=b.dataset.tab;
      document.querySelectorAll('.view').forEach(v=>v.classList.remove('active'));
      document.getElementById(id).classList.add('active');
    });
  });

  // Only-conflicts filter
  const onlyConf=document.getElementById('only-conflicts');
  if(onlyConf){
    onlyConf.addEventListener('change',()=>{
      document.querySelectorAll('#view-calendar .ev').forEach(card=>{
        const has=card.dataset.hasConflict==='1';
        card.style.display=onlyConf.checked?(has?'':'none'):'';
      });
    });
  }

  // "Khác…" select helpers
  function setupOther(selectId, otherId, hiddenId){
    const sel=document.getElementById(selectId), other=document.getElementById(otherId), hidden=document.getElementById(hiddenId);
    function sync(){ if(sel.value==='__OTHER__'){ other.style.display=''; hidden.value=other.value.trim(); } else { other.style.display='none'; hidden.value=sel.value; } }
    sel.addEventListener('change',sync); other.addEventListener('input',sync); sync();
  }
  setupOther('fld-category-select','fld-category-other','fld-category');
  setupOther('fld-location-select','fld-location-other','fld-location');

  // Select or "Khác…"
  function setSelectOrOther(selectId,otherId,hiddenId,value){
    const sel=document.getElementById(selectId), other=document.getElementById(otherId), hidden=document.getElementById(hiddenId);
    const exists=Array.from(sel.options).some(o=>o.value===value);
    if(exists){ sel.value=value; other.style.display='none'; hidden.value=value; }
    else{ sel.value='__OTHER__'; other.style.display=''; other.value=value||''; hidden.value=other.value; }
  }

  // Normalize for attendee compare
  function normLabel(s){ return (s||'').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g,'').replace(/[\\.\\_\\-\\s]+/g,'').toUpperCase().trim(); }

  // Fill form (shared)
  function fillForm(ds){
    const g=k=>(ds[k]??'').toString().trim();
    document.getElementById('fld-id').value=g('id');
    document.getElementById('fld-date').value=g('date');
    document.getElementById('fld-buoi').value=g('buoi');
    document.getElementById('fld-start').value=g('start');
    document.getElementById('fld-end').value=g('end');
    document.getElementById('fld-title').value=g('title');
    (function setSelect(id,val){
      const sel=document.getElementById(id); if(!sel) return;
      const has=Array.from(sel.options).some(o=>o.value===val);
      sel.value=has?val:(sel.value||'');
    })('fld-chair',g('chair'));
    const set=new Set(g('attendees').split(/[,;]+/).map(s=>s.trim()).filter(Boolean).map(normLabel));
    document.querySelectorAll('input[name="attendees"]').forEach(cb=>cb.checked=set.has(normLabel(cb.value)));
    setSelectOrOther('fld-category-select','fld-category-other','fld-category',g('category'));
    setSelectOrOther('fld-location-select','fld-location-other','fld-location',g('location'));
    window.scrollTo({top:0,behavior:'smooth'});
  }

  function editEvent(btn){ const tr=btn.closest('tr'); fillForm(tr.dataset); }
  function editEventFromCard(btn){ const card=btn.closest('.ev'); fillForm(card.dataset); }
</script>
</body>
</html>
"""


# ========== MAIN ==========
if __name__ == "__main__":
    ensure_data_file()
    app.run(debug=True, host="0.0.0.0", port=5000)